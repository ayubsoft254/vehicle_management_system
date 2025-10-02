"""
Reports App - Report File Generators
Generates PDF, Excel, CSV, JSON, and HTML report files
"""

from django.conf import settings
from django.template.loader import render_to_string
from django.utils import timezone
from datetime import datetime
import os
import csv
import json
import logging
from decimal import Decimal

logger = logging.getLogger(__name__)


# ============================================================================
# MAIN GENERATOR FUNCTION
# ============================================================================

def generate_report_file(report, data, output_format, date_from, date_to):
    """
    Generate report file in specified format
    
    Args:
        report: Report instance
        data: Report data dictionary
        output_format: Output format (pdf, excel, csv, json, html)
        date_from: Start date
        date_to: End date
    
    Returns:
        tuple: (file_path, file_size)
    """
    
    generators = {
        'pdf': generate_pdf_report,
        'excel': generate_excel_report,
        'csv': generate_csv_report,
        'json': generate_json_report,
        'html': generate_html_report,
    }
    
    generator = generators.get(output_format)
    
    if not generator:
        raise ValueError(f"Unsupported output format: {output_format}")
    
    return generator(report, data, date_from, date_to)


# ============================================================================
# PDF GENERATOR
# ============================================================================

def generate_pdf_report(report, data, date_from, date_to):
    """
    Generate PDF report
    
    Returns:
        tuple: (file_path, file_size)
    """
    
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.platypus import Image as RLImage
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    except ImportError:
        logger.error("reportlab not installed. Install with: pip install reportlab")
        raise
    
    # Create file path
    file_path = get_report_file_path(report, 'pdf')
    ensure_directory_exists(file_path)
    
    # Create PDF
    doc = SimpleDocTemplate(file_path, pagesize=letter)
    story = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    story.append(Paragraph(report.name, title_style))
    story.append(Spacer(1, 0.2 * inch))
    
    # Report Info
    info_style = styles['Normal']
    report_info = f"""
    <b>Report Type:</b> {report.get_report_type_display()}<br/>
    <b>Period:</b> {date_from.strftime('%B %d, %Y')} - {date_to.strftime('%B %d, %Y')}<br/>
    <b>Generated:</b> {timezone.now().strftime('%B %d, %Y at %H:%M:%S')}<br/>
    """
    story.append(Paragraph(report_info, info_style))
    story.append(Spacer(1, 0.3 * inch))
    
    # Summary Section
    if report.include_summary and 'summary' in data:
        story.append(Paragraph("<b>Summary</b>", styles['Heading2']))
        story.append(Spacer(1, 0.1 * inch))
        
        summary_data = []
        for key, value in data['summary'].items():
            formatted_key = key.replace('_', ' ').title()
            formatted_value = format_value(value)
            summary_data.append([formatted_key, formatted_value])
        
        if summary_data:
            summary_table = Table(summary_data, colWidths=[3 * inch, 2 * inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            story.append(summary_table)
            story.append(Spacer(1, 0.3 * inch))
    
    # Details Section
    if report.include_details:
        # Add tables for different data sections
        for section_name, section_data in data.items():
            if section_name == 'summary':
                continue
            
            if isinstance(section_data, list) and section_data:
                story.append(Paragraph(f"<b>{section_name.replace('_', ' ').title()}</b>", styles['Heading2']))
                story.append(Spacer(1, 0.1 * inch))
                
                # Create table from list data
                table_data = create_table_from_list(section_data)
                if table_data:
                    table = Table(table_data)
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                    ]))
                    story.append(table)
                    story.append(Spacer(1, 0.3 * inch))
    
    # Build PDF
    doc.build(story)
    
    file_size = os.path.getsize(file_path)
    
    return file_path, file_size


# ============================================================================
# EXCEL GENERATOR
# ============================================================================

def generate_excel_report(report, data, date_from, date_to):
    """
    Generate Excel report
    
    Returns:
        tuple: (file_path, file_size)
    """
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl not installed. Install with: pip install openpyxl")
        raise
    
    # Create file path
    file_path = get_report_file_path(report, 'xlsx')
    ensure_directory_exists(file_path)
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Summary Sheet
    if report.include_summary and 'summary' in data:
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Title
        ws_summary['A1'] = report.name
        ws_summary['A1'].font = Font(size=16, bold=True)
        ws_summary['A1'].alignment = Alignment(horizontal='center')
        ws_summary.merge_cells('A1:B1')
        
        # Report info
        ws_summary['A3'] = "Report Type:"
        ws_summary['B3'] = report.get_report_type_display()
        ws_summary['A4'] = "Period:"
        ws_summary['B4'] = f"{date_from.strftime('%Y-%m-%d')} to {date_to.strftime('%Y-%m-%d')}"
        ws_summary['A5'] = "Generated:"
        ws_summary['B5'] = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Summary data
        row = 7
        ws_summary[f'A{row}'] = "Metric"
        ws_summary[f'B{row}'] = "Value"
        ws_summary[f'A{row}'].font = Font(bold=True)
        ws_summary[f'B{row}'].font = Font(bold=True)
        
        row += 1
        for key, value in data['summary'].items():
            ws_summary[f'A{row}'] = key.replace('_', ' ').title()
            ws_summary[f'B{row}'] = format_value(value)
            row += 1
        
        # Style columns
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 20
    
    # Detail Sheets
    if report.include_details:
        for section_name, section_data in data.items():
            if section_name == 'summary':
                continue
            
            if isinstance(section_data, list) and section_data:
                ws = wb.create_sheet(title=section_name[:31])  # Excel limit
                
                # Add headers
                if section_data:
                    headers = list(section_data[0].keys())
                    for col_num, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col_num)
                        cell.value = header.replace('_', ' ').title()
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
                        cell.alignment = Alignment(horizontal='center')
                    
                    # Add data
                    for row_num, item in enumerate(section_data, 2):
                        for col_num, header in enumerate(headers, 1):
                            value = item.get(header)
                            ws.cell(row=row_num, column=col_num, value=format_value(value))
                    
                    # Auto-size columns
                    for col_num in range(1, len(headers) + 1):
                        ws.column_dimensions[get_column_letter(col_num)].width = 15
    
    # Save workbook
    wb.save(file_path)
    
    file_size = os.path.getsize(file_path)
    
    return file_path, file_size


# ============================================================================
# CSV GENERATOR
# ============================================================================

def generate_csv_report(report, data, date_from, date_to):
    """
    Generate CSV report
    
    Returns:
        tuple: (file_path, file_size)
    """
    
    # Create file path
    file_path = get_report_file_path(report, 'csv')
    ensure_directory_exists(file_path)
    
    with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        
        # Header
        writer.writerow([report.name])
        writer.writerow([f"Period: {date_from} to {date_to}"])
        writer.writerow([f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        writer.writerow([])
        
        # Summary
        if report.include_summary and 'summary' in data:
            writer.writerow(['SUMMARY'])
            writer.writerow(['Metric', 'Value'])
            for key, value in data['summary'].items():
                writer.writerow([key.replace('_', ' ').title(), format_value(value)])
            writer.writerow([])
        
        # Details
        if report.include_details:
            for section_name, section_data in data.items():
                if section_name == 'summary':
                    continue
                
                if isinstance(section_data, list) and section_data:
                    writer.writerow([section_name.replace('_', ' ').title().upper()])
                    
                    # Headers
                    headers = list(section_data[0].keys())
                    writer.writerow([h.replace('_', ' ').title() for h in headers])
                    
                    # Data
                    for item in section_data:
                        row = [format_value(item.get(h)) for h in headers]
                        writer.writerow(row)
                    
                    writer.writerow([])
    
    file_size = os.path.getsize(file_path)
    
    return file_path, file_size


# ============================================================================
# JSON GENERATOR
# ============================================================================

def generate_json_report(report, data, date_from, date_to):
    """
    Generate JSON report
    
    Returns:
        tuple: (file_path, file_size)
    """
    
    # Create file path
    file_path = get_report_file_path(report, 'json')
    ensure_directory_exists(file_path)
    
    # Prepare report data
    report_data = {
        'report': {
            'name': report.name,
            'type': report.report_type,
            'period': {
                'from': date_from.isoformat(),
                'to': date_to.isoformat()
            },
            'generated': timezone.now().isoformat()
        },
        'data': convert_to_json_serializable(data)
    }
    
    # Write JSON file
    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(report_data, f, indent=2, ensure_ascii=False)
    
    file_size = os.path.getsize(file_path)
    
    return file_path, file_size


# ============================================================================
# HTML GENERATOR
# ============================================================================

def generate_html_report(report, data, date_from, date_to):
    """
    Generate HTML report
    
    Returns:
        tuple: (file_path, file_size)
    """
    
    # Create file path
    file_path = get_report_file_path(report, 'html')
    ensure_directory_exists(file_path)
    
    # Render HTML template
    context = {
        'report': report,
        'data': data,
        'date_from': date_from,
        'date_to': date_to,
        'generated_at': timezone.now(),
    }
    
    html_content = render_to_string('reports/templates/report_template.html', context)
    
    # Write HTML file
    with open(file_path, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    file_size = os.path.getsize(file_path)
    
    return file_path, file_size


# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def get_report_file_path(report, extension):
    """Generate file path for report"""
    
    # Create reports directory
    reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
    
    # Create subdirectory by date
    date_dir = timezone.now().strftime('%Y/%m')
    full_dir = os.path.join(reports_dir, date_dir)
    
    # Generate filename
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{report.name}_{timestamp}.{extension}"
    
    # Sanitize filename
    filename = sanitize_filename(filename)
    
    return os.path.join(full_dir, filename)


def ensure_directory_exists(file_path):
    """Ensure directory exists for file path"""
    
    directory = os.path.dirname(file_path)
    if not os.path.exists(directory):
        os.makedirs(directory)


def sanitize_filename(filename):
    """Sanitize filename for safe file system use"""
    
    import re
    # Remove invalid characters
    filename = re.sub(r'[^\w\s.-]', '', filename)
    # Replace spaces with underscores
    filename = re.sub(r'\s+', '_', filename)
    return filename


def format_value(value):
    """Format value for display"""
    
    if isinstance(value, (int, float, Decimal)):
        if isinstance(value, Decimal):
            value = float(value)
        # Check if it's a currency value (has 2 decimal places)
        if value != int(value):
            return f"${value:,.2f}" if value > 1000 else f"{value:.2f}"
        return f"{value:,}"
    
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d %H:%M:%S')
    
    if value is None:
        return 'N/A'
    
    return str(value)


def convert_to_json_serializable(obj):
    """Convert object to JSON serializable format"""
    
    if isinstance(obj, dict):
        return {k: convert_to_json_serializable(v) for k, v in obj.items()}
    
    if isinstance(obj, list):
        return [convert_to_json_serializable(item) for item in obj]
    
    if isinstance(obj, Decimal):
        return float(obj)
    
    if isinstance(obj, datetime):
        return obj.isoformat()
    
    return obj


def create_table_from_list(data_list):
    """Create table data from list of dictionaries"""
    
    if not data_list:
        return []
    
    # Get headers from first item
    headers = list(data_list[0].keys())
    formatted_headers = [h.replace('_', ' ').title() for h in headers]
    
    # Create table data
    table_data = [formatted_headers]
    
    for item in data_list[:50]:  # Limit to first 50 rows for PDF
        row = [format_value(item.get(h)) for h in headers]
        table_data.append(row)
    
    return table_data


# ============================================================================
# EXPORT FUNCTIONS
# ============================================================================

def export_report_data(execution, export_format='csv'):
    """
    Export report execution data to specified format
    
    Args:
        execution: ReportExecution instance
        export_format: Export format
    
    Returns:
        str: File path
    """
    
    # Get report data from execution
    data = execution.result_data
    
    if export_format == 'csv':
        return export_to_csv(execution, data)
    elif export_format == 'json':
        return export_to_json(execution, data)
    elif export_format == 'excel':
        return export_to_excel(execution, data)
    
    raise ValueError(f"Unsupported export format: {export_format}")


def export_to_csv(execution, data):
    """Export execution data to CSV"""
    
    file_path = get_report_file_path(execution.report, 'csv')
    ensure_directory_exists(file_path)
    
    with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        
        # Write data
        if isinstance(data, dict):
            for key, value in data.items():
                writer.writerow([key, value])
    
    return file_path


def export_to_json(execution, data):
    """Export execution data to JSON"""
    
    file_path = get_report_file_path(execution.report, 'json')
    ensure_directory_exists(file_path)
    
    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(convert_to_json_serializable(data), f, indent=2)
    
    return file_path


def export_to_excel(execution, data):
    """Export execution data to Excel"""
    
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl required for Excel export")
    
    file_path = get_report_file_path(execution.report, 'xlsx')
    ensure_directory_exists(file_path)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    
    # Write data
    if isinstance(data, dict):
        row = 1
        for key, value in data.items():
            ws.cell(row=row, column=1, value=str(key))
            ws.cell(row=row, column=2, value=format_value(value))
            row += 1
    
    wb.save(file_path)
    
    return file_path