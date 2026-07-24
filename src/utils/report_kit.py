"""
Shared building blocks for every business report in the system.

Every report — Financial, Client Ledger, Business Loans, Expenses, Payroll,
Insurance, Broker/Agent ledgers, Main Ledger, Auctions, Documents, Audit — is
expected to render as: filter form -> on-page preview -> PDF / Excel / CSV
export, using these helpers so all exports share one letterhead, one color
palette and one table style rather than each view inventing its own.
"""
import csv
import io
from decimal import Decimal
from xml.sax.saxutils import escape as _xml_escape

from django.http import HttpResponse
from django.utils import timezone

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape as _landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from utils.letterhead import (
    draw_letterhead, FOOTER_RESERVED_HEIGHT, HEADER_RESERVED_HEIGHT,
)

# ---------------------------------------------------------------------------
# Shared palette — muted slate/navy, deliberately avoids bright gradients
# ---------------------------------------------------------------------------
NAVY = colors.HexColor('#1e293b')
ACCENT = colors.HexColor('#334155')
LIGHT_GREY = colors.HexColor('#f1f5f9')
BORDER_GREY = colors.HexColor('#cbd5e1')
MUTED_TEXT = colors.HexColor('#64748b')

EXCEL_HEADER_FILL = '334155'
EXCEL_HEADER_FONT_COLOR = 'FFFFFF'
EXCEL_ZEBRA_FILL = 'F1F5F9'
EXCEL_BORDER_COLOR = 'CBD5E1'


def fmt_money(amount, currency='KES'):
    amount = amount or 0
    return f"{currency} {amount:,.2f}"


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

def _report_styles():
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        'ReportCompanyName', parent=styles['Normal'], fontSize=14,
        fontName='Helvetica-Bold', textColor=NAVY, leading=16,
    ))
    styles.add(ParagraphStyle(
        'ReportCompanyMeta', parent=styles['Normal'], fontSize=8,
        textColor=MUTED_TEXT, leading=11,
    ))
    styles.add(ParagraphStyle(
        'ReportTitle', parent=styles['Title'], fontSize=16,
        textColor=NAVY, alignment=TA_CENTER, spaceBefore=6, spaceAfter=2,
    ))
    styles.add(ParagraphStyle(
        'ReportSubtitle', parent=styles['Normal'], fontSize=9,
        textColor=MUTED_TEXT, alignment=TA_CENTER, spaceAfter=10,
    ))
    styles.add(ParagraphStyle(
        'ReportSectionHeading', parent=styles['Heading2'], fontSize=11,
        textColor=NAVY, spaceBefore=12, spaceAfter=6,
    ))
    return styles


def build_pdf_response(filename, title, subtitle=None, meta_lines=None,
                        build_body=None, landscape_mode=False):
    """
    Render a complete letterhead + title + footer PDF and return it as a
    file-download HttpResponse. `build_body(elements, styles)` appends the
    report-specific Flowables (KPI blocks, tables, notes). The company
    letterhead (header band + footer trust strip) is drawn on the canvas by
    utils.letterhead so it's identical across every report and doesn't
    consume space in the flowable body.
    """
    buffer = io.BytesIO()
    pagesize = _landscape(A4) if landscape_mode else A4
    doc = SimpleDocTemplate(
        buffer, pagesize=pagesize,
        topMargin=HEADER_RESERVED_HEIGHT + 0.1 * inch,
        bottomMargin=FOOTER_RESERVED_HEIGHT + 0.1 * inch,
        leftMargin=0.75 * inch, rightMargin=0.75 * inch,
    )
    styles = _report_styles()
    elements = []

    elements.append(Paragraph(title.upper(), styles['ReportTitle']))
    sub_bits = [subtitle] if subtitle else []
    sub_bits.append(f"Generated {timezone.now().strftime('%d %B %Y, %H:%M')}")
    elements.append(Paragraph(' &middot; '.join(sub_bits), styles['ReportSubtitle']))

    if meta_lines:
        for line in meta_lines:
            elements.append(Paragraph(line, styles['Normal']))
        elements.append(Spacer(1, 8))

    if build_body:
        build_body(elements, styles)

    doc.build(elements, onFirstPage=draw_letterhead, onLaterPages=draw_letterhead)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def styled_table(data, col_widths=None, align_right_from=None, font_size=8, header=True):
    """A Table flowable using the shared slate header + zebra-striped body."""
    table = Table(data, colWidths=col_widths, repeatRows=1 if header else 0)
    style = [
        ('FONTSIZE', (0, 0), (-1, -1), font_size),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('GRID', (0, 0), (-1, -1), 0.4, BORDER_GREY),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]
    if header:
        style += [
            ('BACKGROUND', (0, 0), (-1, 0), ACCENT),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT_GREY]),
        ]
    if align_right_from is not None:
        style.append(('ALIGN', (align_right_from, 0), (-1, -1), 'RIGHT'))
    table.setStyle(TableStyle(style))
    return table


def kpi_table(pairs, col_widths=None):
    """A compact label/value summary block (bold label, right-aligned value)."""
    data = [[label, value] for label, value in pairs]
    table = Table(data, colWidths=col_widths or [2.6 * inch, 2.6 * inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), LIGHT_GREY),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.4, BORDER_GREY),
    ]))
    return table


# ---------------------------------------------------------------------------
# Width-fitted tables — a wide ledger table must never run off the page.
# Columns are sized proportionally to their content within the printable
# width and every cell wraps, so the fix applies to any column count.
# ---------------------------------------------------------------------------

_MAX_FIT_CHARS = 40   # cap one long cell's influence when apportioning width
_CELL_PADDING_PTS = 12


def _column_content_lengths(headers, rows):
    lengths = []
    for i, header in enumerate(headers):
        max_len = len(str(header))
        for row in rows:
            if i < len(row):
                max_len = max(max_len, len(str(row[i])))
        lengths.append(max_len)
    return lengths


def estimate_table_width(headers, rows, font_size=7.5):
    """Approximate natural (unwrapped) table width in points, used to decide
    whether a report needs landscape orientation."""
    char_width = font_size * 0.55  # average Helvetica glyph width
    return sum(
        min(length, _MAX_FIT_CHARS) * char_width + _CELL_PADDING_PTS
        for length in _column_content_lengths(headers, rows)
    )


def _fit_col_widths(headers, rows, usable_width):
    lengths = [
        min(max(length, 4), _MAX_FIT_CHARS)
        for length in _column_content_lengths(headers, rows)
    ]
    total = sum(lengths) or 1
    widths = [max(0.45 * inch, usable_width * length / total) for length in lengths]
    scale = usable_width / sum(widths)
    return [w * scale for w in widths]


def fitted_table(headers, rows, usable_width, currency_cols=None, totals_row=None, font_size=7.5):
    """
    A styled_table variant guaranteed to fit `usable_width`: cells are
    Paragraphs (so long text wraps instead of being cut off) and column
    widths are apportioned by content. `currency_cols` (1-based, matching
    build_excel_response) right-aligns those columns and renders numeric
    values with thousands separators. `totals_row` renders as a bold
    grand-total row pinned to the bottom of the table.
    """
    currency_cols = currency_cols or set()
    base = ParagraphStyle(
        'FitCell', fontName='Helvetica', fontSize=font_size, leading=font_size + 2.5)
    right = ParagraphStyle('FitCellRight', parent=base, alignment=TA_RIGHT)
    head = ParagraphStyle(
        'FitHead', parent=base, fontName='Helvetica-Bold', textColor=colors.white)
    bold = ParagraphStyle('FitBold', parent=base, fontName='Helvetica-Bold')
    bold_right = ParagraphStyle('FitBoldRight', parent=bold, alignment=TA_RIGHT)

    def cell(value, col, header=False, total=False):
        is_currency = (col + 1) in currency_cols
        if is_currency and isinstance(value, (int, float, Decimal)) and not isinstance(value, bool):
            text = f"{value:,.2f}"
        else:
            text = _xml_escape(str('' if value is None else value))
        if header:
            style = head
        elif total:
            style = bold_right if is_currency else bold
        else:
            style = right if is_currency else base
        return Paragraph(text, style)

    data = [[cell(h, i, header=True) for i, h in enumerate(headers)]]
    for row in rows:
        data.append([cell(v, i) for i, v in enumerate(row)])
    if totals_row:
        data.append([cell(v, i, total=True) for i, v in enumerate(totals_row)])

    width_rows = list(rows) + ([totals_row] if totals_row else [])
    table = Table(
        data,
        colWidths=_fit_col_widths(headers, width_rows, usable_width),
        repeatRows=1,
    )
    style = [
        ('GRID', (0, 0), (-1, -1), 0.4, BORDER_GREY),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('BACKGROUND', (0, 0), (-1, 0), ACCENT),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT_GREY]),
    ]
    if totals_row:
        style += [
            ('BACKGROUND', (0, -1), (-1, -1), BORDER_GREY),
            ('LINEABOVE', (0, -1), (-1, -1), 1, ACCENT),
        ]
    table.setStyle(TableStyle(style))
    return table


def ledger_statement_pdf_response(filename, title, subtitle, statement_rows, statement_summary,
                                   debit_hint=None, credit_hint=None, meta_lines=None):
    """
    A single-entity debit/credit statement PDF - the printable counterpart
    to includes/ledger_statement.html. Used by every party ledger detail
    page (broker, tracker agent, clearing agent, Japan supplier, insurance
    agent) so their "Print" button downloads the same statement shown on
    screen instead of just opening the browser print dialog.
    """
    def body(elements, styles):
        if debit_hint or credit_hint:
            hint_bits = []
            if debit_hint:
                hint_bits.append(f'<b>Debit:</b> {debit_hint}')
            if credit_hint:
                hint_bits.append(f'<b>Credit:</b> {credit_hint}')
            elements.append(Paragraph('  &middot;  '.join(hint_bits), styles['ReportCompanyMeta']))
            elements.append(Spacer(1, 6))

        elements.append(kpi_table([
            ('Opening Balance', fmt_money(statement_summary['opening_balance'])),
            ('Total Debits', fmt_money(statement_summary['total_debits'])),
            ('Total Credits', fmt_money(statement_summary['total_credits'])),
            ('Closing Balance', fmt_money(statement_summary['closing_balance'])),
        ], col_widths=[2.6 * inch, 2.6 * inch]))
        elements.append(Spacer(1, 10))

        headers = ['Date', 'Reference', 'Description', 'Related', 'Method', 'Debit', 'Credit', 'Balance', 'Status']
        table_rows = []
        for row in statement_rows:
            table_rows.append([
                row['date'].strftime('%d %b %Y') if row['date'] else '',
                row['reference'] or '—',
                (row['description'] or '') + (' (REVERSED)' if row.get('is_reversed') else ''),
                row['related'] or '—',
                row['method'] or '—',
                fmt_money(row['debit']) if row['debit'] else '—',
                fmt_money(row['credit']) if row['credit'] else '—',
                fmt_money(row['running_balance']),
                row['status'] or '—',
            ])
        totals_row = [
            '', '', 'GRAND TOTAL', '', '',
            fmt_money(statement_summary['total_debits']),
            fmt_money(statement_summary['total_credits']),
            fmt_money(statement_summary['closing_balance']),
            '',
        ]
        usable_width = _landscape(A4)[0] - 1.5 * inch
        elements.append(fitted_table(
            headers, table_rows, usable_width,
            currency_cols={6, 7, 8}, totals_row=totals_row,
        ))

    return build_pdf_response(
        filename, title, subtitle=subtitle, meta_lines=meta_lines,
        build_body=body, landscape_mode=True,
    )


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def build_excel_response(filename, sheet_title, headers, rows, currency_cols=None, widths=None,
                          totals_row=None):
    """
    A single-sheet workbook with a bold white-on-slate frozen header row,
    zebra striping, right-aligned KES-formatted currency columns, thin
    borders and auto column widths. `totals_row` (same length as headers)
    renders as a bold grand-total row under the data.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    header_font = Font(bold=True, color=EXCEL_HEADER_FONT_COLOR)
    header_fill = PatternFill('solid', fgColor=EXCEL_HEADER_FILL)
    zebra_fill = PatternFill('solid', fgColor=EXCEL_ZEBRA_FILL)
    thin = Side(style='thin', color=EXCEL_BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    ws.freeze_panes = 'A2'

    currency_cols = currency_cols or set()
    for r_idx, row in enumerate(rows, start=2):
        ws.append(row)
        for c_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.border = border
            if r_idx % 2 == 0:
                cell.fill = zebra_fill
            if c_idx in currency_cols:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')

    if totals_row:
        total_fill = PatternFill('solid', fgColor='E2E8F0')
        r_idx = len(rows) + 2
        ws.append(list(totals_row))
        for c_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = Font(bold=True)
            cell.fill = total_fill
            cell.border = border
            if c_idx in currency_cols:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')

    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    else:
        for i, h in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(i)].width = max(12, len(str(h)) + 4)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------

def build_csv_response(filename, headers, rows, totals_row=None):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    if totals_row:
        writer.writerow(totals_row)
    return response


# ---------------------------------------------------------------------------
# Generic single-table dispatcher — for simple reports that are just one
# headers/rows table, so the view only has to build the data once.
# ---------------------------------------------------------------------------

def export_rows(fmt, filename_base, title, headers, rows, currency_cols=None, subtitle=None,
                totals_row=None, landscape_mode=None):
    """
    Render `rows` as a PDF, Excel or CSV download depending on `fmt`.

    `totals_row` (same length as headers) becomes a bold grand-total row at
    the bottom of every format. The PDF fits its table to the printable
    width — switching to landscape automatically when the content is too
    wide for portrait (pass `landscape_mode` to force either way) — and
    wraps long cells rather than letting the table run off the page.
    """
    currency_cols = currency_cols or set()
    if fmt == 'excel':
        return build_excel_response(f'{filename_base}.xlsx', title, headers, rows,
                                    currency_cols=currency_cols, totals_row=totals_row)
    if fmt == 'csv':
        return build_csv_response(f'{filename_base}.csv', headers, rows, totals_row=totals_row)

    all_rows = list(rows) + ([totals_row] if totals_row else [])
    if landscape_mode is None:
        landscape_mode = estimate_table_width(headers, all_rows) > (A4[0] - 1.5 * inch)
    page_width = _landscape(A4)[0] if landscape_mode else A4[0]
    usable_width = page_width - 1.5 * inch

    def body(elements, styles):
        elements.append(fitted_table(
            headers, rows, usable_width,
            currency_cols=currency_cols, totals_row=totals_row,
        ))

    return build_pdf_response(f'{filename_base}.pdf', title, subtitle=subtitle,
                              build_body=body, landscape_mode=landscape_mode)
